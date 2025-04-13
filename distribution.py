# distribution.py
# pip install openpyxl


import tkinter as tk
from tkinter import ttk, messagebox

# Import the database connection functions from db.py
from db import connect_db # Assuming you have a db.py file with the connect_db function defined

import openpyxl
from openpyxl.utils import get_column_letter

def fetch_citizens():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT citizen_id, Name, national_id FROM Citizens")
    rows = cursor.fetchall()
    conn.close()
    return rows

def fetch_goods():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT good_id, name, stock_quantity FROM Goods")
    rows = cursor.fetchall()
    conn.close()
    return rows

def fetch_companies():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT company_id, name FROM Companies")
    rows = cursor.fetchall()
    conn.close()
    return rows

# the following function will be used to add a new distribution record to the database
# and update the stock quantity of the good in the Goods table
def add_distribution(citizen_id, good_id, company_id, quantity, delivery_date):
    conn = connect_db()
    cursor = conn.cursor()

    # Check current stock
    cursor.execute("SELECT stock_quantity FROM Goods WHERE good_id = %s", (good_id,))
    current_stock = cursor.fetchone()[0]

    if quantity > current_stock:
        messagebox.showwarning("Insufficient Stock", "Not enough stock available for this good.")
        conn.close()
        return

    try: # Insert distribution record to the Distributions table in the database
        cursor.execute("""
            INSERT INTO Distributions (citizen_id, good_id, company_id, quantity_given, distribution_date)
            VALUES (%s, %s, %s, %s, %s)
        """, (citizen_id, good_id, company_id, quantity, delivery_date))

        # Decrease stock
        cursor.execute("UPDATE Goods SET stock_quantity = stock_quantity - %s WHERE good_id = %s", (quantity, good_id))

        conn.commit()
        messagebox.showinfo("Success", "Distribution recorded successfully.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to record distribution:\n{e}")
    finally:
        conn.close()


# --- Show Per-Citizen History View ---
def show_per_citizen_history(citizen_id, tree):
    # Fetch distribution history for the specific citizen
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT d.distribution_id, c.Name, g.name, cp.name, d.quantity_given, d.distribution_date
        FROM Distributions d
        JOIN Citizens c ON d.citizen_id = c.citizen_id
        JOIN Goods g ON d.good_id = g.good_id
        JOIN Companies cp ON d.company_id = cp.company_id
        WHERE d.citizen_id = %s
        ORDER BY d.distribution_date DESC
    """, (citizen_id,))
    rows = cursor.fetchall()
    conn.close()

    # Update treeview with citizen's distribution history
    for row in tree.get_children():
        tree.delete(row)
    for dist in rows:
        tree.insert("", "end", values=dist)


# ............................ Window for Distribution Management ............................ UI
 
 
# Function to open the distribution window and manage distributions 
def open_distribution_window():
    window = tk.Toplevel()
    window.title("Distribute Goods to Citizens")
    window.geometry("900x800")

    # Fetch citizens, goods, and companies from the database to populate the dropdowns 
    citizens = fetch_citizens()
    goods = fetch_goods()
    companies = fetch_companies()

    # --- Export Button ---  
    # This button will export the distribution history to an Excel file 
    # It will call the export_to_excel function when clicked
    export_button = tk.Button(window, text="Export to Excel", command=export_to_excel)
    # The button will be placed at the top of the window
    export_button.grid(row=0, column=0, padx=10, pady=10) # Add the export button to the grid



    
    # --- Add Ditribution ---
    # This section is for adding a new Ditribution record to the database 
    # It includes a form for entering the citizen's name, good, quantity received, delivery date, and company
        # Create a frame for the form:
    delivery_frame = tk.LabelFrame(window, text="Record New Ditribution to citizen", padx=10, pady=10)
    delivery_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=5, sticky="nsew")  # Use grid instead of pack

    # bottom for the form 
    tk.Label(delivery_frame, text="Citizen Name:").grid(row=0, column=0, padx=5, pady=5)
    # tk.Label(delivery_frame, text="Citizen ID:").grid(row=0, column=1, padx=5, pady=5) # this not implemented in database yet
    tk.Label(delivery_frame, text="Good:").grid(row=1, column=0, padx=5, pady=5)
    tk.Label(delivery_frame, text="Quantity Received:").grid(row=2, column=0, padx=5, pady=5)
    # tk.Label(delivery_frame, text="Scale: Kg").grid(row=3, column=1, padx=5, pady=5) # this not implemented in database yet
    tk.Label(delivery_frame, text="Delivery Date:").grid(row=3, column=0, padx=5, pady=5) 
    tk.Label(delivery_frame, text="Company:").grid(row=4, column=0, padx=5, pady=5)
    # tk.Label(delivery_frame, text="Delivery ID:").grid(row=6, column=0, padx=5, pady=5) # this not implemented in database yet

    # # Create comboboxes for citizen, good, and company selection 
    # quantity_entry = tk.Entry(form_frame) # Entry for quantity given
    # quantity_entry.grid(row=3, column=1, pady=5) # Add the quantity entry to the grid

    good_var = tk.StringVar()
    citizen_var = tk.StringVar()
    company_var = tk.StringVar()
    quantity_var = tk.StringVar() 
    delivery_date_var = tk.StringVar() # Assuming you have a date picker or entry for the delivery date    


    # quantity_var.insert(0, "Enter Quantity") # Set default text to prompt user
    # quantity_var.bind("<FocusIn>", lambda e: quantity_entry.delete(0, 'end')) # Clear default text on focus

    # Map citizens to their IDs for display 
    goods_map = {f"{g[1]} (Stock: {g[2]})": g[0] for g in goods}
    citizen_map = {f"{c[1]} (ID: {c[2]})": c[0] for c in citizens} # Map citizens to their IDs for display
    company_map = {c[1]: c[0] for c in companies} # Map companies to their IDs for display

    # citizen_menu.grid(row=0, column=1, pady=5, sticky="w")
    # good_menu.grid(row=1, column=1, pady=5, sticky="w")
    # company_menu.grid(row=5, column=1, pady=5, sticky="w")
    # quantity_entry.grid(row=2, column=1, pady=5, sticky="w")
    # citizen_menu.set("Select Citizen") # Set default text to prompt user
    # good_menu.set("Select Good") # Set default text to prompt user
    # company_menu.set("Select Company") # Set default text to prompt user
    # quantity_entry.insert(0, "Enter Quantity") # Set default text to prompt user
    # citizen_menu.bind("<FocusIn>", lambda e: citizen_menu.delete(0, 'end')) # Clear default text on focus
    # good_menu.bind("<FocusIn>", lambda e: good_menu.delete(0, 'end')) # Clear default text on focus
    # company_menu.bind("<FocusIn>", lambda e: company_menu.delete(0, 'end')) # Clear default text on focus
    # quantity_entry.bind("<FocusIn>", lambda e: quantity_entry.delete(0, 'end')) # Clear default text on focus
    # citizen_menu.bind("<FocusOut>", lambda e: citizen_menu.insert(0, "Select Citizen") if not citizen_var.get() else None) # Reset default text on focus out
    # good_menu.bind("<FocusOut>", lambda e: good_menu.insert(0, "Select Good") if not good_var.get() else None) # Reset default text on focus out
    # company_menu.bind("<FocusOut>", lambda e: company_menu.insert(0, "Select Company") if not company_var.get() else None) # Reset default text on focus out
    # quantity_entry.bind("<FocusOut>", lambda e: quantity_entry.insert(0, "Enter Quantity") if not quantity_entry.get() else None) # Reset default text on focus out 

    # Create comboboxes for citizen, good, and company selection
    goods_menu = ttk.Combobox(delivery_frame, textvariable=good_var, values=list(goods_map.keys()), width=40)
    citize_menu = ttk.Combobox(delivery_frame, textvariable=citizen_var, values=list(citizen_map.keys()), width=40)
    company_menu = ttk.Combobox(delivery_frame, textvariable=company_var, values=list(company_map.keys()), width=40)
    quantity_entry = tk.Entry(delivery_frame, textvariable=quantity_var, width=40) # Entry for quantity received 
    delivery_date_entry = tk.Entry(delivery_frame, textvariable=delivery_date_var, width=40) # Assuming you have a date picker or entry for the delivery date

    
    # these are the default values for the comboboxes
    # goods_menu.set("Select Good") # Set default text to prompt user
    # the following lines of grid function will set the position of the widgets in the grid 
    citize_menu.grid(row=0, column=1, pady=5, sticky="w")
    goods_menu.grid(row=1, column=1, pady=5, sticky="w")
    quantity_entry.grid(row=2, column=1, pady=5, sticky="w")
    delivery_date_entry.grid(row=3, column=1, pady=5, sticky="w") # Add the delivery date entry to the grid
    company_menu.grid(row=4, column=1, pady=5, sticky="w")    

    #...........................

    def submit_distribution(): 
        try:
            citizen_id = citizen_map[citizen_var.get()] # this give error if the citizen is not selected 
            good_id = goods_map[good_var.get()]
            company_id = company_map[company_var.get()]
            quantity = int(quantity_entry.get())
            delivery_date = delivery_date_entry.get()  # Assuming you have a date picker or entry for the delivery date

            if quantity <= 0:
                messagebox.showwarning("Invalid Quantity", "Quantity must be greater than 0.")
                return

            add_distribution(citizen_id, good_id, company_id, quantity, delivery_date)
            refresh_distributions()
            quantity_entry.delete(0, 'end')

        except Exception as e:
            messagebox.showerror("Error", f"Check your input:\n{e}")

    # # tk.Button(form_frame, text="Distribute", command=submit_distribution).grid(row=4, column=1, pady=10)
    tk.Button(delivery_frame, text="Submit Distribute", command=submit_distribution).grid(row=5, column=1, pady=20) # Add the distribute button to the grid
    # # tk.Button(form_frame, text="Clear", command=clear_form).grid(row=4, column=2, pady=10) # Add the clear button to the grid

    # # --- Citizen Selection ---
    # citizen_label = tk.Label(window, text="Select Citizen to View History:")
    # citizen_label.pack(pady=5)

    # citizens = fetch_citizens()
    # citizen_map = {f"{c[1]} (ID: {c[2]})": c[0] for c in citizens}
    # citizen_var = tk.StringVar()

    # citizen_menu = ttk.Combobox(window, textvariable=citizen_var, values=list(citizen_map.keys()), width=50)
    # citizen_menu.pack(pady=5)
    # citizen_menu.set("Select Citizen") # Set default text to prompt user

    # the following function will be called when the button is clicked
    # It will fetch the citizen ID from the selected citizen name and call the show_per_citizen_history function
    def show_history_citizen():
        citizen_id = citizen_map[citizen_var.get()]
        show_per_citizen_history(citizen_id, tree)  # Pass 'tree' here to update the treeview
        # # Clear the selection after showing history
        # citizen_var.set("")  # Clear the selection in the combobox 
    # .......................

    
    # citizen_button.pack(pady=10)

    # --- Distribution History Table ---
    history_frame = tk.LabelFrame(window, text="Distribution History:", padx=10, pady=10)
    # history_frame.pack(fill="both", expand=True, padx=10, pady=5)
    history_frame.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky="nsew")

    columns = ("ID", "Citizen", "Good", "Company", "Quantity", "Date")
    tree = ttk.Treeview(history_frame, columns=columns, show="headings")

    # for col in columns:
    #     tree.heading(col, text=col)
    #     tree.column(col, width=130)
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center", width=130)

    # tree.pack(fill="both", expand=True)

    tree.grid(row=0, column=0, sticky="nsew")
    # Allow resizing
    history_frame.grid_rowconfigure(0, weight=1)
    history_frame.grid_columnconfigure(0, weight=1)
    window.grid_rowconfigure(2, weight=1)
    window.grid_columnconfigure(0, weight=1)

    # tree.bind("<Double-1>", lambda e: messagebox.showinfo("Info", "Double-click to view details."))
    
    # This function fetches the distribution history from the database and populates the treeview
    def fetch_distributions(): 
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT d.distribution_id, c.Name, g.name, cp.name, d.quantity_given, d.distribution_date
            FROM Distributions d
            JOIN Citizens c ON d.citizen_id = c.citizen_id
            JOIN Goods g ON d.good_id = g.good_id
            JOIN Companies cp ON d.company_id = cp.company_id
            ORDER BY d.distribution_date DESC
        """)
        rows = cursor.fetchall()
        conn.close()
        return rows
    

    def refresh_distributions():
        for row in tree.get_children():
            tree.delete(row)
        for dist in fetch_distributions():
            tree.insert("", "end", values=dist)

    refresh_distributions() # Call the function to populate the treeview with distribution history

    # # Show the distribution history when the window is opened
    # refresh_distributions() # Call the function to populate the treeview with distribution history

    # --- Distribution History Table ---
    # history_frame = tk.LabelFrame(window, text="Distribution History:", padx=10, pady=10)
    # history_frame.pack(fill="both", expand=True, padx=10, pady=5)
    tk.Button(delivery_frame, text="Show History of Citizen:", command=show_history_citizen).grid(row=6, column=1, pady=10)
    tk.Button(delivery_frame, text="Show All History", command=refresh_distributions).grid(row=6, column=3, pady=10)

    




# --- Function to export the history to Excel ---

def export_to_excel():
    try:
        # Create a new workbook and sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Distribution History"

        # Define the columns
        columns = ("ID", "Citizen", "Good", "Company", "Quantity", "Date")
        
        # Add headers to the sheet
        for col_num, col_name in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}1'] = col_name

        # Fetch data from the database
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT d.distribution_id, c.Name, g.name, cp.name, d.quantity_given, d.distribution_date
            FROM Distributions d
            JOIN Citizens c ON d.citizen_id = c.citizen_id
            JOIN Goods g ON d.good_id = g.good_id
            JOIN Companies cp ON d.company_id = cp.company_id
            ORDER BY d.distribution_date DESC
        """)
        rows = cursor.fetchall()
        conn.close()

        # Write the data to the Excel file
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                col_letter = get_column_letter(col_num)
                sheet[f'{col_letter}{row_num}'] = value

        # Save the workbook to a file
        file_name = "distribution_history.xlsx"
        wb.save(file_name)

        messagebox.showinfo("Success", f"History exported to {file_name}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to export history:\n{e}")


# # Uncomment the following line to run the distribution window directly for testing
# we will considder this as a module and not run it directly
# we will run all in the main.py file
