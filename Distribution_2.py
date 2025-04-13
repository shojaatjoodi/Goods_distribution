# distribution.py
# pip install openpyxl


import tkinter as tk
from tkinter import ttk, messagebox

# Import the database connection functions from db.py
from db import connect_db # Assuming you have a db.py file with the connect_db function defined

import openpyxl
from openpyxl.utils import get_column_letter

def fetch_citizens():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT citizen_id, full_name, national_id FROM Citizens")
    rows = cursor.fetchall()
    conn.close()
    return rows

def fetch_goods():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT good_id, name, stock_quantity FROM Goods")
    rows = cursor.fetchall()
    conn.close()
    return rows

def fetch_companies():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT company_id, name FROM Companies")
    rows = cursor.fetchall()
    conn.close()
    return rows

def add_distribution(citizen_id, good_id, company_id, quantity):
    conn = connect_db()
    cursor = conn.cursor()

    # Check current stock
    cursor.execute("SELECT stock_quantity FROM Goods WHERE good_id = %s", (good_id,))
    current_stock = cursor.fetchone()[0]

    if quantity > current_stock:
        messagebox.showwarning("Insufficient Stock", "Not enough stock available for this good.")
        conn.close()
        return

    try:
        cursor.execute("""
            INSERT INTO Distributions (citizen_id, good_id, company_id, quantity_given)
            VALUES (%s, %s, %s, %s)
        """, (citizen_id, good_id, company_id, quantity))

        # Decrease stock
        cursor.execute("UPDATE Goods SET stock_quantity = stock_quantity - %s WHERE good_id = %s", (quantity, good_id))

        conn.commit()
        messagebox.showinfo("Success", "Distribution recorded successfully.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to record distribution:\n{e}")
    finally:
        conn.close()


# --- Show Per-Citizen History View ---
def show_per_citizen_history(citizen_id, tree):
    # Fetch distribution history for the specific citizen
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT d.distribution_id, c.full_name, g.name, cp.name, d.quantity_given, d.distribution_date
        FROM Distributions d
        JOIN Citizens c ON d.citizen_id = c.citizen_id
        JOIN Goods g ON d.good_id = g.good_id
        JOIN Companies cp ON d.company_id = cp.company_id
        WHERE d.citizen_id = %s
        ORDER BY d.distribution_date DESC
    """, (citizen_id,))
    rows = cursor.fetchall()
    conn.close()

    # Update treeview with citizen's distribution history
    for row in tree.get_children():
        tree.delete(row)
    for dist in rows:
        tree.insert("", "end", values=dist)

# ............................ 
 
# Function to open the distribution window and manage distributions 
def open_distribution_window():
    window = tk.Toplevel()
    window.title("Distribute Goods to Citizens")
    window.geometry("900x800")

    # --- Distribution Form ---
    form_frame = tk.LabelFrame(window, text="New Distribution", padx=10, pady=10)
    form_frame.pack(fill="x", padx=10, pady=5)

    tk.Label(form_frame, text="Citizen:").grid(row=0, column=0, padx=5, pady=5)
    tk.Label(form_frame, text="Good:").grid(row=1, column=0, padx=5, pady=5)
    tk.Label(form_frame, text="Company:").grid(row=2, column=0, padx=5, pady=5)
    tk.Label(form_frame, text="Quantity Given:").grid(row=3, column=0, padx=5, pady=5)

    citizens = fetch_citizens()
    goods = fetch_goods()
    companies = fetch_companies()

    citizen_map = {f"{c[1]} (ID: {c[2]})": c[0] for c in citizens}
    # Map goods to their IDs and stock quantities for display 
    good_map = {f"{g[1]} (Stock: {g[2]})": g[0] for g in goods}
    company_map = {c[1]: c[0] for c in companies}

    citizen_var = tk.StringVar()
    good_var = tk.StringVar()
    company_var = tk.StringVar()
    quantity_entry = tk.Entry(form_frame)

    ttk.Combobox(form_frame, textvariable=citizen_var, values=list(citizen_map.keys()), width=50).grid(row=0, column=1, pady=5)
    ttk.Combobox(form_frame, textvariable=good_var, values=list(good_map.keys()), width=50).grid(row=1, column=1, pady=5)
    ttk.Combobox(form_frame, textvariable=company_var, values=list(company_map.keys()), width=50).grid(row=2, column=1, pady=5)
    quantity_entry.grid(row=3, column=1, pady=5)

    # --- Export Button ---
    export_button = tk.Button(window, text="Export to Excel", command=export_to_excel)
    export_button.pack(pady=10)

    def submit_distribution(): 
        try:
            citizen_id = citizen_map[citizen_var.get()]
            good_id = good_map[good_var.get()]
            company_id = company_map[company_var.get()]
            quantity = int(quantity_entry.get())

            if quantity <= 0:
                messagebox.showwarning("Invalid Quantity", "Quantity must be greater than 0.")
                return

            add_distribution(citizen_id, good_id, company_id, quantity)
            refresh_distributions()
            quantity_entry.delete(0, 'end')

        except Exception as e:
            messagebox.showerror("Error", f"Check your input:\n{e}")

    tk.Button(form_frame, text="Distribute", command=submit_distribution).grid(row=4, column=1, pady=10)


    # --- Citizen Selection ---
    citizen_label = tk.Label(window, text="Select Citizen to View History:")
    citizen_label.pack(pady=5)

    citizens = fetch_citizens()
    citizen_map = {f"{c[1]} (ID: {c[2]})": c[0] for c in citizens}
    citizen_var = tk.StringVar()

    citizen_menu = ttk.Combobox(window, textvariable=citizen_var, values=list(citizen_map.keys()), width=50)
    citizen_menu.pack(pady=5)
    # citizen_menu.set("Select Citizen") # Set default text to prompt user

    # the following function will be called when the button is clicked
    # It will fetch the citizen ID from the selected citizen name and call the show_per_citizen_history function
    def show_history():
        citizen_id = citizen_map[citizen_var.get()]
        show_per_citizen_history(citizen_id, tree)  # Pass 'tree' here to update the treeview
        # # Clear the selection after showing history
        # citizen_var.set("")  # Clear the selection in the combobox 

    citizen_button = tk.Button(window, text="Show History", command=show_history)
    citizen_button.pack(pady=10)

    # --- Distribution History Table ---
    history_frame = tk.LabelFrame(window, text="Distribution History", padx=10, pady=10)
    history_frame.pack(fill="both", expand=True, padx=10, pady=5)

    columns = ("ID", "Citizen", "Good", "Company", "Quantity", "Date")
    tree = ttk.Treeview(history_frame, columns=columns, show="headings")

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=130)

    tree.pack(fill="both", expand=True)
    # tree.bind("<Double-1>", lambda e: messagebox.showinfo("Info", "Double-click to view details."))
    
    # This function fetches the distribution history from the database and populates the treeview
    def fetch_distributions(): 
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT d.distribution_id, c.full_name, g.name, cp.name, d.quantity_given, d.distribution_date
            FROM Distributions d
            JOIN Citizens c ON d.citizen_id = c.citizen_id
            JOIN Goods g ON d.good_id = g.good_id
            JOIN Companies cp ON d.company_id = cp.company_id
            ORDER BY d.distribution_date DESC
        """)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def refresh_distributions():
        for row in tree.get_children():
            tree.delete(row)
        for dist in fetch_distributions():
            tree.insert("", "end", values=dist)

    refresh_distributions()




# --- Function to export the history to Excel ---

def export_to_excel():
    try:
        # Create a new workbook and sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Distribution History"

        # Define the columns
        columns = ("ID", "Citizen", "Good", "Company", "Quantity", "Date")
        
        # Add headers to the sheet
        for col_num, col_name in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}1'] = col_name

        # Fetch data from the database
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT d.distribution_id, c.full_name, g.name, cp.name, d.quantity_given, d.distribution_date
            FROM Distributions d
            JOIN Citizens c ON d.citizen_id = c.citizen_id
            JOIN Goods g ON d.good_id = g.good_id
            JOIN Companies cp ON d.company_id = cp.company_id
            ORDER BY d.distribution_date DESC
        """)
        rows = cursor.fetchall()
        conn.close()

        # Write the data to the Excel file
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                col_letter = get_column_letter(col_num)
                sheet[f'{col_letter}{row_num}'] = value

        # Save the workbook to a file
        file_name = "distribution_history.xlsx"
        wb.save(file_name)

        messagebox.showinfo("Success", f"History exported to {file_name}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to export history:\n{e}")


# # Uncomment the following line to run the distribution window directly for testing
# we will considder this as a module and not run it directly
# we will run all in the main.py file
